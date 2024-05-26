
import torch
import numpy as np
import pickle    #可进行序列化操作
import math
import random
import matplotlib.pyplot as plt
import seaborn as sns
from scipy.stats import norm
from xlutils.copy import copy
from openpyxl import load_workbook,Workbook
import xlrd
import xlwt
sns.set(context='notebook', font='simhei', style='whitegrid')
# 设置风格尺度和显示中文
import warnings
warnings.filterwarnings('ignore')  # 不发出警告
plt.rcParams['font.sans-serif']=['SimHei'] #用来正常显示中文标签
plt.rcParams['axes.unicode_minus']=False #用来正常显示负号
book_name_xls = 'E:/pruning/kernel_pruning/ori.xls'
sheet_name_xls = '1'
def write_excel_xls(path, sheet_name, value):

    workbook = xlwt.Workbook()  # 新建一个工作簿
    sheet = workbook.add_sheet(sheet_name)  # 在工作簿中新建一个表格
    for i in range(0, 1):
        for j in range(0, 3):
            sheet.write(i, j, value[j])  # 像表格中写入数据（对应的行和列）
    workbook.save(path)  # 保存工作簿
    print("xls格式表格写入数据成功！")


def write_excel_xls_append(path, value):

    workbook = xlrd.open_workbook(path)  # 打开工作簿
    sheets = workbook.sheet_names()  # 获取工作簿中的所有表格
    worksheet = workbook.sheet_by_name(sheets[0])  # 获取工作簿中所有表格中的的第一个表格
    rows_old = worksheet.nrows  # 获取表格中已存在的数据的行数
    new_workbook = copy(workbook)  # 将xlrd对象拷贝转化为xlwt对象
    new_worksheet = new_workbook.get_sheet(0)  # 获取转化后工作簿中的第一个表格
    for i in range(0, 1):
        for j in range(0, 3):
            new_worksheet.write(i + rows_old, j, value[j])  # 追加写入数据，注意是从i+rows_old行开始写入
    new_workbook.save(path)  # 保存工作簿
    print("xls格式表格【追加】写入数据成功！")


def read_excel_xls(path):
    workbook = xlrd.open_workbook(path)  # 打开工作簿
    sheets = workbook.sheet_names()  # 获取工作簿中的所有表格
    worksheet = workbook.sheet_by_name(sheets[0])  # 获取工作簿中所有表格中的的第一个表格
    for i in range(0, worksheet.nrows):
        for j in range(0, worksheet.ncols):
            print(worksheet.cell_value(i, j), "\t", end="")  # 逐行逐列读取数据
        print()
def statistics(s,covid):
    s = s.flatten()
    zero_dir = np.where(s == 0)
    s = np.delete(s , zero_dir )
    plt.figure(figsize=(8, 6),num=covid)
    sns.distplot(s, bins=14, hist=True, kde=False, norm_hist=False,
                 rug=True, vertical=False, label='Density distribution',rug_kws={'color':'r'},
                 axlabel='Matrix elements sum',hist_kws={'color': 'b', 'edgecolor': 'k'},fit=norm)
    font1 = {'family': 'Times New Roman','weight': 'normal','size': 22,}
    plt.legend(prop=font1)
    plt.tick_params(labelsize=16)
    font2 = {'family': 'Times New Roman','weight': 'normal', 'size': 22,}
    plt.xlabel('Matrix elements sum', font2)
    plt.grid(linestyle='--')
    plt.savefig('./fig/cov{}.png'.format(covid),dpi=300)
    plt.show()
    plt.close()
    # VAL = []
    # s0 = (s>=-0.1) & (s<=0.1)
    # num0 = np.sum( s0 ==True ) /float(len(s))
    # P = np.sum(s>0) /float(len(s))
    # N = np.sum(s<0) /float(len(s))
    # VAL.append(num0)
    # VAL.append(P)
    # VAL.append(N)
    # if covid == 1:
    #     write_excel_xls(book_name_xls, sheet_name_xls, VAL)
    # else:
    #     write_excel_xls_append(book_name_xls, VAL)

class mask_vgg_16_bn:
    def __init__(self, model=None, compress_rate=[0.50], job_dir='',device=None):
        self.model = model
        self.compress_rate = compress_rate
        self.cpra = []
        self.mask = {}
        self.job_dir=job_dir
        self.device = device
        self.ws = torch.tensor(0)
        self.bn = torch.tensor(0)
    def layer_mask(self, cov_id, resume=None, param_per_cov=4,  arch="vgg_16_bn", way='A'):
        params = self.model.parameters()
        params = list(params)

        if resume:
            with open(resume, 'rb') as f:
                self.mask = pickle.load(f)  # 反序列化对象。将文件中的数据解析为一个Python对象。
        else:           # ./result/tmp/
            resume=self.job_dir+'/mask'     # 有直接读取这个mask, 没有 resume ，下面会写入得到新的mask文件

        self.param_per_cov=param_per_cov

        if way == 'A' :
            for index, item in enumerate(params):
                    #卷积核的权重和还是需要将权重和bias共同加起来
                if index == (cov_id - 1) * param_per_cov:
                    f, c, w, h = item.size()
                    self.ws = torch.tensor([torch.sum(item[i, j, :, :]) for i in range(f) for j in range(c)])
                    self.ws = self.ws.view(f, -1).float()
                    self.ws = np.array(self.ws)
                    # bs = torch.tensor(0)
                    bs = torch.as_tensor(params[index+1])
                    bs = bs.detach().cpu().numpy()
                    for i in range(len(self.ws)):
                        for j in range(len(self.ws[0])):
                            self.ws[i][j] = self.ws[i][j] + bs[i]
                    statistics(self.ws,cov_id)
                    self.bn = torch.as_tensor(params[index+2])
                    self.bn = self.bn.detach().cpu().numpy()
                    # 一个滤波器64个卷积核 每个卷积核3个通道 （64，3） 只是将3个里面的某些通道置0，所以无需对bias  BN层进行操作
                    #固定剪枝率，从重要性最低的滤波器开始除草，并计数
                    # self.ws = np.round((self.ws /np.max(self.ws)) * (5-3*(cov_id/12))  )       #  self.ws:权重和
                    self.ws = np.round( ( 1/(  1+np.exp(-(self.ws/np.max(self.ws)))  )-0.5 )*20 ) # (1/(1+e^-x)-0.5)*10
                    ind = np.argsort(abs(self.bn))[:]                                   # ind：滤波器重要性的索引 作用：从小到大排序。
                    pruned_num = int(self.compress_rate[cov_id - 1] * f*c)
                    ones_i = torch.ones(f, c).to(self.device)                           # ones_i:掩码 作用：对要删除的卷积核的位置清0

                    for i in range(len(ind)):                                           # 作用：从重要性最低的滤波器开始剪枝
                        ws_value_list = list(set(self.ws[ind[i]]))                      # ws_value_list:无序列表 作用：存放不重复的数
                        for x in range(len(ws_value_list) - 1):                         # 外层循环确定比较的轮数，x是下标，lt[x]在外层循环中代表lt中所有元素
                            for y in range(x + 1, len(ws_value_list)):
                                if np.sum(self.ws == ws_value_list[x]) < np.sum(self.ws == ws_value_list[y]):
                                    ws_value_list[x],ws_value_list[y] = ws_value_list[y],ws_value_list[x]
                        for j in range(len(ws_value_list)) :
                            if pruned_num == 0:
                                break
                            num = np.sum(self.ws[ind[i]] == ws_value_list[j])           #当前冗余值的个数
                            val_index = np.where(self.ws[ind[i]] == ws_value_list[j] )  #当前冗余值所在的索引值
                            index_1 = list(range(0,num))
                            index_1 = random.sample(index_1,math.ceil(num*(1-np.max(self.compress_rate)-0.1)))
                            val_index = np.delete(val_index, index_1, axis=1)       #将要保留的索引删除
                            for m in range(len(val_index[0])) :                     #对剩余的索引进行位置置0
                                ones_i[ind[i]][val_index[0][m]] = 0
                                pruned_num -= 1
                                if pruned_num == 0:
                                    break
                        if pruned_num == 0:
                            break
                    cnt_array = np.sum(ones_i.cpu().numpy() == 0)
                    self.cpra.append(format(cnt_array / (f * c), '.2g'))
                    ones = torch.ones(f, c, w, h).to(self.device)
                    for i in range(f):
                        for j in range(c):
                            for k in range(w):
                                for l in range(h):
                                    ones[i, j, k, l] = ones_i[i, j]
                    self.mask[index] = ones
                    item.data = item.data * self.mask[index]

                    with open(resume, "wb") as f:  # 保存每一层的mask 0,1
                        pickle.dump(self.mask, f)
                    break
        elif way == 'B' :
            for index, item in enumerate(params):
                if index == (cov_id - 1) * param_per_cov:
                    f, c, w, h = item.size()
                    item = item.to(self.device)
                    self.ws = torch.tensor([torch.sum(item[i, j, :, :]) for i in range(f) for j in range(c)])
                    self.ws = np.array(self.ws.view(f, -1).float())
                    # 一个滤波器64个卷积核 每个卷积核3个通道 （64，3） 只是将3个里面的某些通道置0，所以无需对bias  BN层进行操作
                    #固定剪枝率，从重要性最低的滤波器开始除草，并计数
                    self.ws = np.round((self.ws /np.max(self.ws)) * 5) # self.ws:权重和
                    # self.ws = np.round((1 / (1 + np.exp(-(self.ws / np.max(self.ws)))) - 0.5) * 5)  # (1/(1+e^-x)-0.5)*10
                    ws = self.ws[0]
                    for n in range(f - 1) :
                        ws = np.hstack((ws, self.ws[n+1]))
                    pruned_num = int(self.compress_rate[cov_id - 1] * f*c)
                    ones_i = torch.ones(f, c).to(self.device)                           # ones_i:掩码 作用：对要删除的卷积核的位置清0
                    ws_value_list = list(set(ws))                      # ws_value_list:无序列表 作用：存放不重复的数
                    for x in range(len(ws_value_list) - 1):            # 循环比较，从大到小排列
                        for y in range(x + 1, len(ws_value_list)):
                            if np.sum(self.ws == ws_value_list[x]) < np.sum(self.ws == ws_value_list[y]):
                                ws_value_list[x], ws_value_list[y] = ws_value_list[y], ws_value_list[x]
                    for j in range(len(ws_value_list)) :
                        if pruned_num == 0:
                            break
                        num = np.sum(self.ws == ws_value_list[j])           #当前冗余值的个数
                        val_index = np.where(self.ws == ws_value_list[j] )  #当前冗余值所在的索引值
                        index_1 = list(range(0, num))
                        index_1 = random.sample(index_1, math.ceil(num*(1-np.max(self.compress_rate)-0.1)))
                        val_index = np.delete(val_index, index_1, axis=1)       #将要保留的索引删除
                        for m in range(len(val_index[0])) :                     #对剩余的索引进行位置置0
                            ones_i[val_index[0][m]][val_index[1][m]] = 0
                            pruned_num -= 1
                            if pruned_num == 0:
                                break
                        if pruned_num == 0:
                            break
                    cnt_array = np.sum(ones_i.cpu().numpy() == 0)
                    self.cpra.append(format(cnt_array / (f * c), '.2g'))
                    ones = torch.ones(f, c, w, h).to(self.device)
                    for i in range(f):
                        for j in range(c):
                            for k in range(w):
                                for l in range(h):
                                    ones[i, j, k, l] = ones_i[i, j]
                    self.mask[index] = ones
                    item.data = item.data * self.mask[index]

                    with open(resume, "wb") as f:  # 保存每一层的mask 0,1
                        pickle.dump(self.mask, f)
                    break
        else :
            assert 1 == 0

    def grad_mask(self, cov_id,epoch):   # 掩码和模型参数相乘
        params = self.model.parameters()
        for index, item in enumerate(params):
            if index == cov_id * self.param_per_cov:
                break
            if index in range(0, 48, self.param_per_cov):
                if epoch < 20 :
                    self.mask[index][self.mask[index] != 1] = 1-(epoch+1)*0.05
                item.data = item.data * self.mask[index].to(self.device)

    def pri_comp(self):
        print('compress_rate:{}'.format(self.cpra))

class mask_resnet_56:
    def __init__(self, model=None, compress_rate=[0.50], job_dir='', device=None):
        self.model = model
        self.compress_rate = compress_rate
        self.mask = {}
        self.cpra = []
        self.job_dir = job_dir
        self.device = device
        self.ws = torch.tensor(0)
        self.bn = torch.tensor(0)

    def layer_mask(self, cov_id, resume=None, param_per_cov=3, arch="resnet_56",way='A'):
        params = self.model.parameters()
        params = list(params)

        if resume:
            with open(resume, 'rb') as f:
                self.mask = pickle.load(f)  # 反序列化对象。将文件中的数据解析为一个Python对象。
        else:  # ./result/tmp/
            resume = self.job_dir + '/mask'  # 有直接读取这个mask, 没有 resume ，下面会写入得到新的mask文件

        self.param_per_cov = param_per_cov

        if way == 'A':
            for index, item in enumerate(params):
                # 卷积核的权重和还是需要将权重和bias共同加起来
                if index == (cov_id - 1) * param_per_cov:
                    f, c, w, h = item.size()
                    self.ws = torch.tensor([torch.sum(item[i, j, :, :]) for i in range(f) for j in range(c)])
                    self.ws = self.ws.view(f, -1).float()
                    self.ws = np.array(self.ws)
                    statistics(self.ws,cov_id)
                    self.bn = torch.as_tensor(params[index + 1])
                    self.bn = self.bn.detach().cpu().numpy()
                    # 一个滤波器64个卷积核 每个卷积核3个通道 （64，3） 只是将3个里面的某些通道置0，所以无需对bias  BN层进行操作
                    # 固定剪枝率，从重要性最低的滤波器开始除草，并计数
                    # self.ws = np.round((self.ws / np.max(self.ws)) * (5 - 3 * (cov_id / 12)))  # self.ws:权重和
                    self.ws = np.round((1 / (1 + np.exp(-(self.ws / np.max(self.ws)))) - 0.5) * 20)  # (1/(1+e^-x)-0.5)*10
                    ind = np.argsort(abs(self.bn))[:]  # ind：滤波器重要性的索引 作用：从小到大排序。
                    pruned_num = int(self.compress_rate[cov_id - 1] * f * c)
                    ones_i = torch.ones(f, c).to(self.device)  # ones_i:掩码 作用：对要删除的卷积核的位置清0

                    for i in range(len(ind)):  # 作用：从重要性最低的滤波器开始剪枝
                        ws_value_list = list(set(self.ws[ind[i]]))  # ws_value_list:无序列表 作用：存放不重复的数
                        for x in range(len(ws_value_list) - 1):  # 循环比较，从大到小排列
                            for y in range(x + 1, len(ws_value_list)):
                                if np.sum(self.ws == ws_value_list[x]) < np.sum(self.ws == ws_value_list[y]):
                                    ws_value_list[x], ws_value_list[y] = ws_value_list[y], ws_value_list[x]
                        for j in range(len(ws_value_list)):
                            if pruned_num == 0:
                                break
                            num = np.sum(self.ws[ind[i]] == ws_value_list[j])  # 当前冗余值的个数
                            val_index = np.where(self.ws[ind[i]] == ws_value_list[j])  # 当前冗余值所在的索引值
                            index_1 = list(range(0, num))
                            index_1 = random.sample(index_1, math.ceil(num*(1-np.max(self.compress_rate)-0.1)))
                            val_index = np.delete(val_index, index_1, axis=1)  # 将要保留的索引删除
                            for m in range(len(val_index[0])):  # 对剩余的索引进行位置置0
                                ones_i[ind[i]][val_index[0][m]] = 0
                                pruned_num -= 1
                                if pruned_num == 0:
                                    break
                        if pruned_num == 0:
                            break
                    # pruned_num = int(self.compress_rate[cov_id - 1] * f * c)
                    # ones_i = torch.ones(f, c).to(self.device)  # ones_i:掩码 作用：对要删除的卷积核的位置清0
                    # for i in range(len(self.bn)):  # 作用：从重要性最低的层开始剪枝
                    #     for j in range(1, len(self.ws[i])):
                    #         if pruned_num == 0:
                    #             break
                    #         else:  # 作用：非0的数重复后进行清0
                    #             ones_i[i][j] = 0
                    #             pruned_num -= 1
                    #     if pruned_num == 0:
                    #         break
                    cnt_array = np.sum(ones_i.cpu().numpy() == 0)
                    self.cpra.append(format(cnt_array / (f * c), '.2g'))
                    ones = torch.ones(f, c, w, h).to(self.device)
                    for i in range(f):
                        for j in range(c):
                            for k in range(w):
                                for l in range(h):
                                    ones[i, j, k, l] = ones_i[i, j]
                    self.mask[index] = ones
                    item.data = item.data * self.mask[index]

                    with open(resume, "wb") as f:  # 保存每一层的mask 0,1
                        pickle.dump(self.mask, f)
                    break
        elif way == 'B':
            for index, item in enumerate(params):
                if index == (cov_id - 1) * param_per_cov:
                    f, c, w, h = item.size()
                    self.ws = torch.tensor([torch.sum(item[i, j, :, :]) for i in range(f) for j in range(c)])
                    self.ws = self.ws.view(f, -1).float()
                    self.ws = np.array(self.ws)
                    # 一个滤波器64个卷积核 每个卷积核3个通道 （64，3） 只是将3个里面的某些通道置0，所以无需对bias  BN层进行操作
                    # 固定剪枝率，从重要性最低的滤波器开始除草，并计数
                    self.ws = np.round((self.ws / np.max(self.ws)) * 5)  # self.ws:权重和
                    # self.ws = np.round((1 / (1 + np.exp(-(self.ws / np.max(self.ws)))) - 0.5) * 5)  # (1/(1+e^-x)-0.5)*10
                    ws = self.ws[0]
                    for n in range(f - 1):
                        ws = np.hstack((ws, self.ws[n + 1]))
                    pruned_num = int(self.compress_rate[cov_id - 1] * f * c)
                    ones_i = torch.ones(f, c).to(self.device)  # ones_i:掩码 作用：对要删除的卷积核的位置清0
                    ws_value_list = list(set(ws))  # ws_value_list:无序列表 作用：存放不重复的数
                    for x in range(len(ws_value_list) - 1):            # 循环比较，从大到小排列
                        for y in range(x + 1, len(ws_value_list)):
                            if np.sum(self.ws == ws_value_list[x]) < np.sum(self.ws == ws_value_list[y]):
                                ws_value_list[x], ws_value_list[y] = ws_value_list[y], ws_value_list[x]
                    for j in range(len(ws_value_list)):
                        if pruned_num == 0:
                            break
                        num = np.sum(self.ws == ws_value_list[j])  # 当前冗余值的个数
                        val_index = np.where(self.ws == ws_value_list[j])  # 当前冗余值所在的索引值
                        index_1 = list(range(0, num))
                        index_1 = random.sample(index_1, math.ceil(num*(1-np.max(self.compress_rate)-0.1)))
                        val_index = np.delete(val_index, index_1, axis=1)  # 将要保留的索引删除
                        for m in range(len(val_index[0])):  # 对剩余的索引进行位置置0
                            ones_i[val_index[0][m]][val_index[1][m]] = 0
                            pruned_num -= 1
                            if pruned_num == 0:
                                break
                        if pruned_num == 0:
                            break
                    cnt_array = np.sum(ones_i.cpu().numpy() == 0)
                    self.cpra.append(format(cnt_array / (f * c), '.2g'))
                    ones = torch.ones(f, c, w, h).to(self.device)
                    for i in range(f):
                        for j in range(c):
                            for k in range(w):
                                for l in range(h):
                                    ones[i, j, k, l] = ones_i[i, j]
                    self.mask[index] = ones
                    item.data = item.data * self.mask[index]

                    with open(resume, "wb") as f:  # 保存每一层的mask 0,1
                        pickle.dump(self.mask, f)
                    break
        else:
            assert 1 == 0

    def grad_mask(self, cov_id):
        params = self.model.parameters()
        for index, item in enumerate(params):
            if index == cov_id * self.param_per_cov:
                break
            if index in range(0, 167, self.param_per_cov):
                item.data = item.data * self.mask[index].to(self.device)  # prune certain weight

    def pri_comp(self):
        print('compress_rate:{}'.format(self.cpra))

class mask_googlenet:
    def __init__(self, model=None, compress_rate=[0.50], job_dir='',device=None):
        self.model = model
        self.compress_rate = compress_rate
        self.cpra = []
        self.mask = {}
        self.job_dir=job_dir
        self.device = device
        self.ws = torch.tensor(0)
        self.bn = torch.tensor(0)

    def layer_mask(self, cov_id, resume=None, param_per_cov=28,  arch="googlenet",way='A'):
        params = self.model.parameters()
        params = list(params)
        if resume:
            with open(resume, 'rb') as f:
                self.mask = pickle.load(f)
        else:
            resume=self.job_dir+'/mask'

        self.param_per_cov=param_per_cov

        if way == 'A':
            for index, item in enumerate(params):
                if index == (cov_id-1) * param_per_cov + 4:
                    break
                if (cov_id == 1 and index == 0) \
                        or index == (cov_id - 1) * param_per_cov - 24 \
                        or index == (cov_id - 1) * param_per_cov - 16 \
                        or index == (cov_id - 1) * param_per_cov - 8 \
                        or index == (cov_id - 1) * param_per_cov - 4 \
                        or index == (cov_id - 1) * param_per_cov:
                    f, c, w, h = item.size()
                self.ws = torch.tensor([torch.sum(item[i, j, :, :]) for i in range(f) for j in range(c)])
                self.ws = self.ws.view(f, -1).float()
                self.ws = np.array(self.ws)
                bs = torch.as_tensor(params[index + 1])
                bs = bs.detach().cpu().numpy()
                for i in range(len(self.ws)):
                    for j in range(len(self.ws[0])):
                        self.ws[i][j] = self.ws[i][j] + bs[i]
                # zhifang(self.ws,cov_id)
                self.bn = torch.as_tensor(params[index + 2])
                self.bn = self.bn.detach().cpu().numpy()
                # 一个滤波器64个卷积核 每个卷积核3个通道 （64，3） 只是将3个里面的某些通道置0，所以无需对bias  BN层进行操作
                # 固定剪枝率，从重要性最低的滤波器开始除草，并计数
                self.ws = np.round((self.ws / np.max(self.ws)) * (5 - 3 * (cov_id / 12)))  # self.ws:权重和
                # self.ws = np.round((1 / (1 + np.exp(-(self.ws / np.max(self.ws)))) - 0.5) * 5)  # (1/(1+e^-x)-0.5)*10
                ind = np.argsort(abs(self.bn))[:]  # ind：滤波器重要性的索引 作用：从小到大排序。
                pruned_num = int(self.compress_rate[cov_id - 1] * f * c)
                ones_i = torch.ones(f, c).to(self.device)  # ones_i:掩码 作用：对要删除的卷积核的位置清0

                for i in range(len(ind)):  # 作用：从重要性最低的滤波器开始剪枝
                    ws_value_list = list(set(self.ws[ind[i]]))  # ws_value_list:无序列表 作用：存放不重复的数
                    for x in range(len(ws_value_list) - 1):            # 循环比较，从大到小排列
                        for y in range(x + 1, len(ws_value_list)):
                            if np.sum(self.ws == ws_value_list[x]) < np.sum(self.ws == ws_value_list[y]):
                                ws_value_list[x], ws_value_list[y] = ws_value_list[y], ws_value_list[x]
                    for j in range(len(ws_value_list)):
                        if pruned_num == 0:
                            break
                        num = np.sum(self.ws[ind[i]] == ws_value_list[j])  # 当前冗余值的个数
                        val_index = np.where(self.ws[ind[i]] == ws_value_list[j])  # 当前冗余值所在的索引值
                        index_1 = list(range(0, num))
                        index_1 = random.sample(index_1, math.ceil(num*(1-np.max(self.compress_rate)-0.1)))
                        val_index = np.delete(val_index, index_1, axis=1)  # 将要保留的索引删除
                        for m in range(len(val_index[0])):  # 对剩余的索引进行位置置0
                            ones_i[ind[i]][val_index[0][m]] = 0
                            pruned_num -= 1
                            if pruned_num == 0:
                                break
                    if pruned_num == 0:
                        break
                cnt_array = np.sum(ones_i.cpu().numpy() == 0)
                self.cpra.append(format(cnt_array / (f * c), '.2g'))
                ones = torch.ones(f, c, w, h).to(self.device)
                for i in range(f):
                    for j in range(c):
                        for k in range(w):
                            for l in range(h):
                                ones[i, j, k, l] = ones_i[i, j]
                self.mask[index] = ones
                item.data = item.data * self.mask[index]

                with open(resume, "wb") as f:  # 保存每一层的mask 0,1
                    pickle.dump(self.mask, f)
                break
        elif way == 'B':
            for index, item in enumerate(params):
                if index == (cov_id - 1) * param_per_cov + 4:
                    break
                if (cov_id == 1 and index == 0) \
                        or index == (cov_id - 1) * param_per_cov - 24 \
                        or index == (cov_id - 1) * param_per_cov - 16 \
                        or index == (cov_id - 1) * param_per_cov - 8 \
                        or index == (cov_id - 1) * param_per_cov - 4 \
                        or index == (cov_id - 1) * param_per_cov:
                    f, c, w, h = item.size()
                    self.ws = torch.tensor([torch.sum(item[i, j, :, :]) for i in range(f) for j in range(c)])
                    self.ws = self.ws.view(f, -1).float()
                    self.ws = np.array(self.ws)
                    # 一个滤波器64个卷积核 每个卷积核3个通道 （64，3） 只是将3个里面的某些通道置0，所以无需对bias  BN层进行操作
                    # 固定剪枝率，从重要性最低的滤波器开始除草，并计数
                    self.ws = np.round((self.ws / np.max(self.ws)) * 5)  # self.ws:权重和
                    # self.ws = np.round((1 / (1 + np.exp(-(self.ws / np.max(self.ws)))) - 0.5) * 5)  # (1/(1+e^-x)-0.5)*10
                    ws = self.ws[0]
                    for n in range(f - 1):
                        ws = np.hstack((ws, self.ws[n + 1]))
                    pruned_num = int(self.compress_rate[cov_id - 1] * f * c)
                    ones_i = torch.ones(f, c).to(self.device)  # ones_i:掩码 作用：对要删除的卷积核的位置清0
                    ws_value_list = list(set(ws))  # ws_value_list:无序列表 作用：存放不重复的数
                    for x in range(len(ws_value_list) - 1):            # 循环比较，从大到小排列
                        for y in range(x + 1, len(ws_value_list)):
                            if np.sum(self.ws == ws_value_list[x]) < np.sum(self.ws == ws_value_list[y]):
                                ws_value_list[x], ws_value_list[y] = ws_value_list[y], ws_value_list[x]
                    for j in range(len(ws_value_list)):
                        if pruned_num == 0:
                            break
                        num = np.sum(self.ws == ws_value_list[j])  # 当前冗余值的个数
                        val_index = np.where(self.ws == ws_value_list[j])  # 当前冗余值所在的索引值
                        index_1 = list(range(0, num))
                        index_1 = random.sample(index_1, math.ceil(num*(1-np.max(self.compress_rate)-0.1)))
                        val_index = np.delete(val_index, index_1, axis=1)  # 将要保留的索引删除
                        for m in range(len(val_index[0])):  # 对剩余的索引进行位置置0
                            ones_i[val_index[0][m]][val_index[1][m]] = 0
                            pruned_num -= 1
                            if pruned_num == 0:
                                break
                        if pruned_num == 0:
                            break
                    cnt_array = np.sum(ones_i.cpu().numpy() == 0)
                    self.cpra.append(format(cnt_array / (f * c), '.2g'))
                    ones = torch.ones(f, c, w, h).to(self.device)
                    for i in range(f):
                        for j in range(c):
                            for k in range(w):
                                for l in range(h):
                                    ones[i, j, k, l] = ones_i[i, j]
                    self.mask[index] = ones
                    item.data = item.data * self.mask[index]

                    with open(resume, "wb") as f:  # 保存每一层的mask 0,1
                        pickle.dump(self.mask, f)
                    break
        else:
            assert 1 == 0

    def grad_mask(self, cov_id):
        params = self.model.parameters()
        for index, item in enumerate(params):
            if index == (cov_id-1) * self.param_per_cov + 4:
                break
            if index not in self.mask:
                continue
            item.data = item.data * self.mask[index].to(self.device)#prune certain weight

    def pri_comp(self):
        print('compress_rate:{}'.format(self.cpra))

class mask_resnet_110:
    def __init__(self, model=None, compress_rate=[0.50], job_dir='',device=None):
        self.model = model
        self.compress_rate = compress_rate
        self.cpra = []
        self.mask = {}
        self.job_dir=job_dir
        self.device = device
        self.ws = torch.tensor(0)
        self.bn = torch.tensor(0)
    def layer_mask(self, cov_id, resume=None, param_per_cov=3,  arch="resnet_110_convwise",way='A'):
        params = self.model.parameters()
        params = list(params)
        if resume:
            with open(resume, 'rb') as f:
                self.mask = pickle.load(f)
        else:
            resume=self.job_dir+'/mask'

        self.param_per_cov=param_per_cov

        if way == 'A':
            for index, item in enumerate(params):
                # 卷积核的权重和还是需要将权重和bias共同加起来
                if index == (cov_id - 1) * param_per_cov:
                    f, c, w, h = item.size()
                    self.ws = torch.tensor([torch.sum(item[i, j, :, :]) for i in range(f) for j in range(c)])
                    self.ws = self.ws.view(f, -1).float()
                    self.ws = np.array(self.ws)
                    # zhifang(self.ws,cov_id)
                    self.bn = torch.as_tensor(params[index + 1])
                    self.bn = self.bn.detach().cpu().numpy()
                    # 一个滤波器64个卷积核 每个卷积核3个通道 （64，3） 只是将3个里面的某些通道置0，所以无需对bias  BN层进行操作
                    # 固定剪枝率，从重要性最低的滤波器开始除草，并计数
                    self.ws = np.round((self.ws / np.max(self.ws)) * (5 - 3 * (cov_id / 12)))  # self.ws:权重和
                    # self.ws = np.round((1 / (1 + np.exp(-(self.ws / np.max(self.ws)))) - 0.5) * 5)  # (1/(1+e^-x)-0.5)*10
                    ind = np.argsort(abs(self.bn))[:]  # ind：滤波器重要性的索引 作用：从小到大排序。
                    pruned_num = int(self.compress_rate[cov_id - 1] * f * c)
                    ones_i = torch.ones(f, c).to(self.device)  # ones_i:掩码 作用：对要删除的卷积核的位置清0

                    for i in range(len(ind)):  # 作用：从重要性最低的滤波器开始剪枝
                        ws_value_list = list(set(self.ws[ind[i]]))  # ws_value_list:无序列表 作用：存放不重复的数
                        for x in range(len(ws_value_list) - 1):  # 循环比较，从大到小排列
                            for y in range(x + 1, len(ws_value_list)):
                                if np.sum(self.ws == ws_value_list[x]) < np.sum(self.ws == ws_value_list[y]):
                                    ws_value_list[x], ws_value_list[y] = ws_value_list[y], ws_value_list[x]
                        for j in range(len(ws_value_list)):
                            if pruned_num == 0:
                                break
                            num = np.sum(self.ws[ind[i]] == ws_value_list[j])  # 当前冗余值的个数
                            val_index = np.where(self.ws[ind[i]] == ws_value_list[j])  # 当前冗余值所在的索引值

                            index_1 = list(range(0, num))
                            index_1 = random.sample(index_1, math.ceil(num*(1-np.max(self.compress_rate)-0.1)))
                            val_index = np.delete(val_index, index_1, axis=1)  # 将要保留的索引删除
                            for m in range(len(val_index[0])):  # 对剩余的索引进行位置置0
                                ones_i[ind[i]][val_index[0][m]] = 0
                                pruned_num -= 1
                                if pruned_num == 0:
                                    break
                        if pruned_num == 0:
                            break
                    cnt_array = np.sum(ones_i.cpu().numpy() == 0)
                    self.cpra.append(format(cnt_array / (f * c), '.2g'))
                    ones = torch.ones(f, c, w, h).to(self.device)
                    for i in range(f):
                        for j in range(c):
                            for k in range(w):
                                for l in range(h):
                                    ones[i, j, k, l] = ones_i[i, j]
                    self.mask[index] = ones
                    item.data = item.data * self.mask[index]

                    with open(resume, "wb") as f:  # 保存每一层的mask 0,1
                        pickle.dump(self.mask, f)
                    break
        elif way == 'B':
            for index, item in enumerate(params):
                if index == (cov_id - 1) * param_per_cov:
                    f, c, w, h = item.size()
                    self.ws = torch.tensor([torch.sum(item[i, j, :, :]) for i in range(f) for j in range(c)])
                    self.ws = self.ws.view(f, -1).float()
                    self.ws = np.array(self.ws)
                    # 一个滤波器64个卷积核 每个卷积核3个通道 （64，3） 只是将3个里面的某些通道置0，所以无需对bias  BN层进行操作
                    # 固定剪枝率，从重要性最低的滤波器开始除草，并计数
                    self.ws = np.round((self.ws / np.max(self.ws)) * 5)  # self.ws:权重和
                    self.ws = np.round((1 / (1 + np.exp(-(self.ws / np.max(self.ws)))) - 0.5) * 5)  # (1/(1+e^-x)-0.5)*10
                    ws = self.ws[0]
                    for n in range(f - 1):
                        ws = np.hstack((ws, self.ws[n + 1]))
                    pruned_num = int(self.compress_rate[cov_id - 1] * f * c)
                    ones_i = torch.ones(f, c).to(self.device)  # ones_i:掩码 作用：对要删除的卷积核的位置清0
                    ws_value_list = list(set(ws))  # ws_value_list:无序列表 作用：存放不重复的数
                    for x in range(len(ws_value_list) - 1):            # 循环比较，从大到小排列
                        for y in range(x + 1, len(ws_value_list)):
                            if np.sum(self.ws == ws_value_list[x]) < np.sum(self.ws == ws_value_list[y]):
                                ws_value_list[x], ws_value_list[y] = ws_value_list[y], ws_value_list[x]
                    for j in range(len(ws_value_list)):
                        if pruned_num == 0:
                            break
                        num = np.sum(self.ws == ws_value_list[j])  # 当前冗余值的个数
                        val_index = np.where(self.ws == ws_value_list[j])  # 当前冗余值所在的索引值
                        index_1 = list(range(0, num))
                        index_1 = random.sample(index_1, math.ceil(num*(1-np.max(self.compress_rate)-0.1)))
                        val_index = np.delete(val_index, index_1, axis=1)  # 将要保留的索引删除
                        for m in range(len(val_index[0])):  # 对剩余的索引进行位置置0
                            ones_i[val_index[0][m]][val_index[1][m]] = 0
                            pruned_num -= 1
                            if pruned_num == 0:
                                break
                        if pruned_num == 0:
                            break
                    cnt_array = np.sum(ones_i.cpu().numpy() == 0)
                    self.cpra.append(format(cnt_array / (f * c), '.2g'))
                    ones = torch.ones(f, c, w, h).to(self.device)
                    for i in range(f):
                        for j in range(c):
                            for k in range(w):
                                for l in range(h):
                                    ones[i, j, k, l] = ones_i[i, j]
                    self.mask[index] = ones
                    item.data = item.data * self.mask[index]

                    with open(resume, "wb") as f:  # 保存每一层的mask 0,1
                        pickle.dump(self.mask, f)
                    break
        else:
            assert 1 == 0

    def grad_mask(self, cov_id):
        params = self.model.parameters()
        for index, item in enumerate(params):
            if index == cov_id*self.param_per_cov:
                break
            if index in range(0, 326, self.param_per_cov):
                item.data = item.data * self.mask[index].to(self.device)  # prune certain weight

    def pri_comp(self):
        print('compress_rate:{}'.format(self.cpra))

class mask_resnet_50:
    def __init__(self, model=None, compress_rate=[0.50], job_dir='',device=None):
        self.model = model
        self.compress_rate = compress_rate
        self.mask = {}
        self.cpra = []
        self.job_dir = job_dir
        self.device = device
        self.ws = torch.tensor(0)
        self.bn = torch.tensor(0)

    def layer_mask(self, cov_id, resume=None, param_per_cov=3, arch="resnet_56", way='A'):
        params = self.model.parameters()
        params = list(params)

        if resume:
            with open(resume, 'rb') as f:
                self.mask = pickle.load(f)  # 反序列化对象。将文件中的数据解析为一个Python对象。
        else:  # ./result/tmp/
            resume = self.job_dir + '/mask'  # 有直接读取这个mask, 没有 resume ，下面会写入得到新的mask文件

        self.param_per_cov = param_per_cov

        if way == 'A':
            for index, item in enumerate(params):
                # 卷积核的权重和还是需要将权重和bias共同加起来
                if index == (cov_id - 1) * param_per_cov:
                    f, c, w, h = item.size()
                    self.ws = torch.tensor([torch.sum(item[i, j, :, :]) for i in range(f) for j in range(c)])
                    self.ws = self.ws.view(f, -1).float()
                    self.ws = np.array(self.ws)
                    # zhifang(self.ws,cov_id)
                    self.bn = torch.as_tensor(params[index + 1])
                    self.bn = self.bn.detach().cpu().numpy()
                    # 一个滤波器64个卷积核 每个卷积核3个通道 （64，3） 只是将3个里面的某些通道置0，所以无需对bias  BN层进行操作
                    # 固定剪枝率，从重要性最低的滤波器开始除草，并计数
                    self.ws = np.round((self.ws / np.max(self.ws)) * (5 - 3 * (cov_id / 12)))  # self.ws:权重和
                    # self.ws = np.round((1 / (1 + np.exp(-(self.ws / np.max(self.ws)))) - 0.5) * 5)  # (1/(1+e^-x)-0.5)*10
                    ind = np.argsort(abs(self.bn))[:]  # ind：滤波器重要性的索引 作用：从小到大排序。
                    pruned_num = int(self.compress_rate[cov_id - 1] * f * c)
                    ones_i = torch.ones(f, c).to(self.device)  # ones_i:掩码 作用：对要删除的卷积核的位置清0

                    for i in range(len(ind)):  # 作用：从重要性最低的滤波器开始剪枝
                        ws_value_list = list(set(self.ws[ind[i]]))  # ws_value_list:无序列表 作用：存放不重复的数
                        for x in range(len(ws_value_list) - 1):  # 循环比较，从大到小排列
                            for y in range(x + 1, len(ws_value_list)):
                                if np.sum(self.ws == ws_value_list[x]) < np.sum(self.ws == ws_value_list[y]):
                                    ws_value_list[x], ws_value_list[y] = ws_value_list[y], ws_value_list[x]
                        for j in range(len(ws_value_list)):
                            if pruned_num == 0:
                                break
                            num = np.sum(self.ws[ind[i]] == ws_value_list[j])  # 当前冗余值的个数
                            val_index = np.where(self.ws[ind[i]] == ws_value_list[j])  # 当前冗余值所在的索引值

                            index_1 = list(range(0, num))
                            index_1 = random.sample(index_1, math.ceil(num*(1-np.max(self.compress_rate)-0.1)))
                            val_index = np.delete(val_index, index_1, axis=1)  # 将要保留的索引删除
                            for m in range(len(val_index[0])):  # 对剩余的索引进行位置置0
                                ones_i[ind[i]][val_index[0][m]] = 0
                                pruned_num -= 1
                                if pruned_num == 0:
                                    break
                        if pruned_num == 0:
                            break
                    cnt_array = np.sum(ones_i.cpu().numpy() == 0)
                    self.cpra.append(format(cnt_array / (f * c), '.2g'))
                    ones = torch.ones(f, c, w, h).to(self.device)
                    for i in range(f):
                        for j in range(c):
                            for k in range(w):
                                for l in range(h):
                                    ones[i, j, k, l] = ones_i[i, j]
                    self.mask[index] = ones
                    item.data = item.data * self.mask[index]

                    with open(resume, "wb") as f:  # 保存每一层的mask 0,1
                        pickle.dump(self.mask, f)
                    break
        elif way == 'B':
            for index, item in enumerate(params):
                if index == (cov_id - 1) * param_per_cov:
                    f, c, w, h = item.size()
                    self.ws = torch.tensor([torch.sum(item[i, j, :, :]) for i in range(f) for j in range(c)])
                    self.ws = self.ws.view(f, -1).float()
                    self.ws = np.array(self.ws)
                    # 一个滤波器64个卷积核 每个卷积核3个通道 （64，3） 只是将3个里面的某些通道置0，所以无需对bias  BN层进行操作
                    # 固定剪枝率，从重要性最低的滤波器开始除草，并计数
                    self.ws = np.round((self.ws / np.max(self.ws)) * 5)  # self.ws:权重和
                    # self.ws = np.round((1 / (1 + np.exp(-(self.ws / np.max(self.ws)))) - 0.5) * 5)  # (1/(1+e^-x)-0.5)*10
                    ws = self.ws[0]
                    for n in range(f - 1):
                        ws = np.hstack((ws, self.ws[n + 1]))
                    pruned_num = int(self.compress_rate[cov_id - 1] * f * c)
                    ones_i = torch.ones(f, c).to(self.device)  # ones_i:掩码 作用：对要删除的卷积核的位置清0
                    ws_value_list = list(set(ws))  # ws_value_list:无序列表 作用：存放不重复的数
                    for x in range(len(ws_value_list) - 1):            # 循环比较，从大到小排列
                        for y in range(x + 1, len(ws_value_list)):
                            if np.sum(self.ws == ws_value_list[x]) < np.sum(self.ws == ws_value_list[y]):
                                ws_value_list[x], ws_value_list[y] = ws_value_list[y], ws_value_list[x]
                    for j in range(len(ws_value_list)):
                        if pruned_num == 0:
                            break
                        num = np.sum(self.ws == ws_value_list[j])  # 当前冗余值的个数
                        val_index = np.where(self.ws == ws_value_list[j])  # 当前冗余值所在的索引值
                        index_1 = list(range(0, num))
                        index_1 = random.sample(index_1, math.ceil(num*(1-np.max(self.compress_rate)-0.1)))
                        val_index = np.delete(val_index, index_1, axis=1)  # 将要保留的索引删除
                        for m in range(len(val_index[0])):  # 对剩余的索引进行位置置0
                            ones_i[val_index[0][m]][val_index[1][m]] = 0
                            pruned_num -= 1
                            if pruned_num == 0:
                                break
                        if pruned_num == 0:
                            break
                    cnt_array = np.sum(ones_i.cpu().numpy() == 0)
                    self.cpra.append(format(cnt_array / (f * c), '.2g'))
                    ones = torch.ones(f, c, w, h).to(self.device)
                    for i in range(f):
                        for j in range(c):
                            for k in range(w):
                                for l in range(h):
                                    ones[i, j, k, l] = ones_i[i, j]
                    self.mask[index] = ones
                    item.data = item.data * self.mask[index]

                    with open(resume, "wb") as f:  # 保存每一层的mask 0,1
                        pickle.dump(self.mask, f)
                    break
        else:
            assert 1 == 0

    def grad_mask(self, cov_id):
        params = self.model.parameters()
        for index, item in enumerate(params):
            if index == cov_id * self.param_per_cov:
                break
            if index in range(0, 161, self.param_per_cov):
                item.data = item.data * self.mask[index].to(self.device)  # prune certain weight

    def pri_comp(self):
        print('compress_rate:{}'.format(self.cpra))